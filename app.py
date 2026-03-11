"""Web app: upload Excel timeplan, enter teacher name and code(s), download one unified ICS calendar."""

import io
import os
import tempfile
from pathlib import Path

import openpyxl
from flask import Flask, request, send_file, jsonify

from excel2ics import export_all_classes_of_teacher

app = Flask(__name__, static_folder="static", static_url_path="")
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024  # 32 MB

ALLOWED_EXTENSIONS = {"xlsx", "xls"}


def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


@app.route("/")
def index():
    return send_file(Path(__file__).parent / "static" / "index.html")


@app.route("/api/convert", methods=["POST"])
def convert():
    """Accept Excel file + teacher name + teacher code(s); return one unified ICS."""
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400
    file = request.files["file"]
    if not file or file.filename == "":
        return jsonify({"error": "No file selected"}), 400
    if not allowed_file(file.filename):
        return jsonify({"error": "Only .xlsx and .xls files are allowed"}), 400

    teacher_name = request.form.get("teacher_name", "").strip()
    if not teacher_name:
        return jsonify({"error": "Teacher name is required"}), 400

    teacher_codes_raw = request.form.get("teacher_codes", "").strip()
    if not teacher_codes_raw:
        return jsonify({"error": "Enter at least one teacher code (e.g. RS7, RS4)"}), 400
    teacher_codes = [c.strip() for c in teacher_codes_raw.split(",") if c.strip()]
    if not teacher_codes:
        return jsonify({"error": "Enter at least one teacher code"}), 400

    sheet_name = request.form.get("sheet", "").strip() or "06 Timeplan"
    tmp_path = None
    tmp_ics_path = None

    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            file.save(tmp.name)
            tmp_path = tmp.name

        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return jsonify(
                {
                    "error": f"Sheet '{sheet_name}' not found. Available sheets: {', '.join(wb.sheetnames)}"
                }
            ), 400
        wb.close()

        fd, tmp_ics_path = tempfile.mkstemp(suffix=".ics")
        os.close(fd)

        export_all_classes_of_teacher(
            tmp_path,
            teacher_name=teacher_name,
            teacher_codes=teacher_codes,
            output_path=tmp_ics_path,
            sheet_name=sheet_name,
            invite_emails=False,
        )

        with open(tmp_ics_path, encoding="utf-8") as f:
            ics_content = f.read()
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if tmp_ics_path:
            Path(tmp_ics_path).unlink(missing_ok=True)
        if tmp_path:
            Path(tmp_path).unlink(missing_ok=True)

    safe_name = (
        "".join(c if c.isalnum() else "_" for c in teacher_name.lower()).strip("_") or "calendar"
    )
    filename = f"{safe_name}_calendar.ics"
    # UTF-8 with BOM for Outlook compatibility
    ics_bytes = ics_content.encode("utf-8")
    buf = io.BytesIO(b"\xef\xbb\xbf" + ics_bytes)
    return send_file(
        buf,
        mimetype="text/calendar; charset=utf-8",
        as_attachment=True,
        download_name=filename,
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5050, debug=True)
