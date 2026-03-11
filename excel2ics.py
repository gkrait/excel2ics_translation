"""Extract teacher calendar from the Excel timeplan and export to ICS format.

This module parses the "06 Timeplan" sheet from the building school schedule
and extracts class schedules for specific teachers (RS7, RS4).

Algorithm:
1. Find all cells containing the teacher name (e.g., "RS7", "RS4") in the calendar area
   (columns up to AQ on sheet "06 Timeplan", part "TIMEPLAN - BYGG 2025-2026")
2. For each found cell at (row, col):
   a) Look upward in same column for first cell with group label
      (1BYA, 2BYA, 1BYC, 2BYC, 3BYC, 4BYC, etc.) - this is the student group
   b) Find the date by mapping the column to its day start column (H, N, T, Z, AF, AL)
      and looking upward for the date
   c) Count class rows from date row to current cell to determine time slot:
      - Classes start at 8:00 AM
      - Each class is 90 minutes
      - 15-minute break between classes
      - No classes between 15:00-16:00
   d) Look downward in same column for first cell with room number pattern

Structure:
- Days start at columns: H(8), N(14), T(20), Z(26), AF(32), AL(38)
- Each day has 6 columns
- Column layout per day: [Subject, Teacher, ?, Subject, Teacher, ?] (two parallel tracks)
- Teacher columns are offset +1 from day start (columns 9, 12 for Monday, etc.)
"""

from datetime import datetime, timezone
from pathlib import Path
import re
from typing import NamedTuple

import openpyxl


class ClassSession(NamedTuple):
    """Represents a single class session."""

    date: datetime
    start_time: str
    end_time: str
    subject: str
    teacher_name: str
    teacher_code: str
    student_group: str
    room: str | None


# Column AQ = column 43
MAX_CALENDAR_COLUMN = 43

# Day start columns (where dates are located)
DAY_START_COLUMNS = [8, 14, 20, 26, 32, 38]  # H, N, T, Z, AF, AL

# Student group pattern
GROUP_PATTERN = re.compile(r"^\d+BY[A-Z]$")

# Room number pattern (typically 4 digits)
ROOM_PATTERN = re.compile(r"^\d{4}$")


def get_day_start_column(col: int) -> int:
    """Get the day start column for a given column.

    Each day spans 6 columns. Returns the first column of that day.
    """
    for i, day_start in enumerate(DAY_START_COLUMNS):
        day_end = day_start + 6
        if day_start <= col < day_end:
            return day_start
    # If beyond known days, estimate
    if col >= 38:
        return 38 + ((col - 38) // 6) * 6
    return 8


def is_group_label(value) -> bool:
    """Check if a value is a student group label (1BYA, 2BYA, etc.)."""
    if value is None:
        return False
    if not isinstance(value, str):
        return False
    return bool(GROUP_PATTERN.match(value.strip()))


def is_room_number(value) -> bool:
    """Check if a value looks like a room number."""
    if value is None:
        return False
    value_str = str(value).strip()
    # Room numbers are typically 4 digits
    return bool(ROOM_PATTERN.match(value_str))


def is_class_cell(cell) -> bool:
    """Check if a cell is a class cell (has content that's not a header).

    Class cells contain teacher names, subject codes, etc.
    Header cells contain dates, group labels, "Rom:", "Uke", etc.
    """
    value = cell.value
    if value is None:
        return False

    value_str = str(value).strip()
    if not value_str:
        return False

    # Exclude known header patterns
    if value_str in ("Uke", "Rom:", "None"):
        return False
    if is_group_label(value_str):
        return False
    if isinstance(value, datetime):
        return False

    # If it has content and isn't a header, it's a class cell
    return True


def find_date_for_column(ws, row: int, col: int) -> tuple[datetime, int] | None:
    """Find the date for a given cell by looking at the day start column.

    Searches upward from the given row in the day's start column.

    Args:
        ws: Worksheet
        row: Current row
        col: Current column

    Returns:
        Tuple of (date, date_row) or None if not found
    """
    day_start_col = get_day_start_column(col)

    # Search upward for a date in the day start column
    for r in range(row, 0, -1):
        cell = ws.cell(row=r, column=day_start_col)
        if isinstance(cell.value, datetime):
            return (cell.value, r)
    return None


def find_group_above(ws, row: int, col: int) -> str | None:
    """Find the student group label by searching upward.

    The group label could be:
    1. In the same column
    2. In the start of the current track (left track: offset 0, right track: offset 3)

    Args:
        ws: Worksheet
        row: Starting row (exclusive)
        col: Column to search

    Returns:
        Group label string or None if not found
    """
    day_start_col = get_day_start_column(col)
    offset_in_day = col - day_start_col

    # Determine which track we're in and which column has the group
    # Left track (offset 0-2): group in offset 0
    # Right track (offset 3-5): group in offset 3
    if offset_in_day < 3:
        group_col = day_start_col  # Left track
    else:
        group_col = day_start_col + 3  # Right track

    # Search upward for group name in the track's group column
    for r in range(row - 1, max(0, row - 20), -1):
        cell = ws.cell(row=r, column=group_col)
        if is_group_label(cell.value):
            return str(cell.value).strip()

    # Also check the current column in case group is directly above
    for r in range(row - 1, max(0, row - 20), -1):
        cell = ws.cell(row=r, column=col)
        if is_group_label(cell.value):
            return str(cell.value).strip()

    return None


def find_room_below(ws, row: int, col: int, max_row: int) -> str | None:
    """Find the room number by searching downward for a room number pattern.

    Args:
        ws: Worksheet
        row: Starting row (exclusive)
        col: Column to search
        max_row: Maximum row to search

    Returns:
        Room number string or None if not found
    """
    for r in range(row + 1, min(row + 10, max_row + 1)):
        cell = ws.cell(row=r, column=col)
        if is_room_number(cell.value):
            return str(cell.value).strip()
    return None


def count_class_rows(ws, date_row: int, current_row: int, col: int) -> int:
    """Count the number of class rows from date_row to current_row.

    A class row is a row that has content in the current column
    (teacher names, subject codes, etc., not headers).

    Args:
        ws: Worksheet
        date_row: Row containing the date
        current_row: Current cell row
        col: Column to check

    Returns:
        Count of class rows from date_row+1 to current_row (inclusive)
    """
    count = 0
    for r in range(date_row + 1, current_row + 1):
        cell = ws.cell(row=r, column=col)
        if is_class_cell(cell):
            count += 1
    return count


def calculate_time_slot(
    class_row_count: int, class_date: datetime | None = None
) -> tuple[str, str] | None:
    """Calculate start and end time based on row offset from date row.

    Time slots:
    - Saturday: single slot 9:00-15:00
    - Other days: class_row_count = row - date_row; if <= 6 → morning slot,
      otherwise afternoon slot.

    Args:
        class_row_count: Row difference (row - date_row). <= 6 → morning, else afternoon.
        class_date: Date of the class (used to detect Saturday)

    Returns:
        Tuple of (start_time, end_time) or None if invalid
    """
    # Saturday: single slot 9:00-15:00
    if class_date is not None and class_date.weekday() == 5:
        return ("09:00", "15:00")

    # Time slots before 15:00
    # Based on user example: 3rd slot is 11:45-13:15 (longer break before slot 3)
    morning_slots = [
        ("08:00", "09:30"),  # Slot 1
        ("09:45", "11:15"),  # Slot 2
        ("11:45", "13:15"),  # Slot 3 (30 min break before)
        ("13:30", "15:00"),  # Slot 4
    ]

    # Time slots after 16:00 (no classes 15:00-16:00)
    afternoon_slots = [
        ("16:30", "20:00"),  # Slot 5
    ]

    if class_row_count < 0:
        print(f"Invalid class_row_count: {class_row_count}")
        return None

    # Morning: row offset <= 6; afternoon: > 6. Map 1–6 to morning slots (1–4).
    if class_row_count <= 6:
        slot_index = min(class_row_count, len(morning_slots))
        return morning_slots[slot_index]
    return afternoon_slots[0]


def _merge_sequential_slots(classes: list[ClassSession]) -> list[ClassSession]:
    """Merge sequential time slots into single longer sessions.

    Two classes are considered sequential and merged if:
    - Same date
    - Same subject
    - Same student group
    - Same teacher
    - End time of first is close to start time of second (within 45 minutes)

    Args:
        classes: List of class sessions to merge

    Returns:
        List of merged class sessions
    """
    if not classes:
        return classes

    # Sort by date, subject, group, teacher_code, then start time
    sorted_classes = sorted(
        classes,
        key=lambda c: (c.date.date(), c.subject, c.student_group, c.teacher_code, c.start_time),
    )

    merged: list[ClassSession] = []
    current = sorted_classes[0]

    for next_class in sorted_classes[1:]:
        # Check if this class can be merged with the current one
        same_date = current.date.date() == next_class.date.date()
        same_subject = current.subject == next_class.subject
        same_group = current.student_group == next_class.student_group
        same_teacher = current.teacher_code == next_class.teacher_code

        if same_date and same_subject and same_group and same_teacher:
            # Parse times to check if sequential
            current_end_h, current_end_m = map(int, current.end_time.split(":"))
            next_start_h, next_start_m = map(int, next_class.start_time.split(":"))

            current_end_minutes = current_end_h * 60 + current_end_m
            next_start_minutes = next_start_h * 60 + next_start_m

            # Sequential if gap is <= 45 minutes (accounts for breaks and the 15:00-16:00 gap)
            gap_minutes = next_start_minutes - current_end_minutes

            if 0 <= gap_minutes <= 60:
                # Merge: extend current session to include next_class
                current = ClassSession(
                    date=current.date,
                    start_time=current.start_time,
                    end_time=next_class.end_time,
                    subject=current.subject,
                    teacher_name=current.teacher_name,
                    teacher_code=current.teacher_code,
                    student_group=current.student_group,
                    room=current.room or next_class.room,  # Use first non-None room
                )
                continue

        # Not sequential, save current and move to next
        merged.append(current)
        current = next_class

    # Don't forget the last one
    merged.append(current)

    return merged


def extract_classes_for_teacher(
    excel_path: str,
    teacher_code: str,
    teacher_name: str | None = None,
    sheet_name: str = "06 Timeplan",
) -> list[ClassSession]:
    """Extract all classes for a specific teacher from the timeplan.

    Args:
        excel_path: Path to the Excel file
        teacher_code: String to search for in the Excel (e.g. "RS7", "RS4")
        teacher_name: Name to use in the ICS/calendar; if None, defaults to teacher_code
        sheet_name: Name of the sheet containing the timeplan

    Returns:
        List of ClassSession objects
    """
    if teacher_name is None:
        teacher_name = teacher_code

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[sheet_name]

    classes: list[ClassSession] = []
    seen_classes: set[tuple] = set()  # To deduplicate

    max_row = ws.max_row or 1

    # Step 1: Find all cells containing the teacher code in calendar area
    print(f"Searching for '{teacher_code}' in columns A-AQ...")
    interesting_cells = []
    for row in range(1, max_row + 1):
        for col in range(1, MAX_CALENDAR_COLUMN + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value and teacher_code in str(cell.value):
                interesting_cells.append((row, col, str(cell.value)))

    print(f"Found {len(interesting_cells)} cells with '{teacher_code}'")

    # Deduplicate consecutive rows in same column (merged cells)
    # Keep only the first row of each merged cell group
    print(f"After deduplication: {len(interesting_cells)} unique cells")

    # Step 2: Process each interesting cell
    for row, col, teacher_value in interesting_cells:
        # 2-a) Find student group
        student_group = find_group_above(ws, row, col)

        # 2-b) Find date for this column
        date_info = find_date_for_column(ws, row, col)
        if date_info is None:
            continue
        class_date, date_row = date_info

        # 2-c) Row offset from date row: <= 6 → morning slot, otherwise afternoon
        class_row_count = (
            row - date_row - 2
        )  # -2 because the first two rows are the date and the group
        time_slot = calculate_time_slot(class_row_count, class_date)
        if time_slot is None:
            continue
        start_time, end_time = time_slot

        # 2-d) Find room (search downward)
        room = find_room_below(ws, row, col, max_row)

        # Get subject (one column to the left of teacher)
        subject_cell = ws.cell(row=row, column=col - 1)
        subject = str(subject_cell.value) if subject_cell.value else "Unknown"

        # Create a unique key to avoid duplicates
        class_key = (
            class_date.date(),
            start_time,
            end_time,
            teacher_code,
            subject,
            student_group,
        )
        if class_key in seen_classes:
            continue
        seen_classes.add(class_key)

        classes.append(
            ClassSession(
                date=class_date,
                start_time=start_time,
                end_time=end_time,
                subject=subject,
                teacher_name=teacher_name,
                teacher_code=teacher_code,
                student_group=student_group or "Unknown",
                room=room,
            )
        )

    # Step 3: Merge sequential time slots into one
    # Two classes are sequential if they are on the same date, same subject, same group,
    # and the end_time of one matches (or is close to) the start_time of the next
    classes = _merge_sequential_slots(classes)

    wb.close()

    # Sort by date and time
    classes.sort(key=lambda c: (c.date, c.start_time))

    return classes


def _escape_ics_param(value: str) -> str:
    """Escape a value for use in ICS property parameters (e.g. CN=)."""
    if ";" in value or "," in value or "\\" in value or '"' in value:
        return '"' + value.replace("\\", "\\\\").replace('"', '\\"') + '"'
    return value


def _escape_ics_text(value: str) -> str:
    """Escape text for ICS property values (SUMMARY, DESCRIPTION, etc.)."""
    return value.replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")


def _fold_ics_line(line: str, max_len: int = 75) -> str:
    """Fold a long ICS content line per RFC 5545 (CRLF + space for continuation)."""
    if len(line) <= max_len:
        return line
    folded = [line[:max_len]]
    rest = line[max_len:]
    while rest:
        folded.append(" " + rest[: max_len - 1])
        rest = rest[max_len - 1 :]
    return "\r\n".join(folded)


def _safe_uid(uid: str) -> str:
    """Ensure UID only contains characters safe for Outlook (no spaces, minimal set)."""
    return "".join(c for c in uid if c.isalnum() or c in "-._@")


def classes_to_ics(
    classes: list[ClassSession],
    teacher_name: str,
    teacher_code: str | None = None,
    method: str = "METHOD:PUBLISH",
    invite_emails: bool = False,
    organizer_email: str | None = None,
) -> str:
    """Convert a list of ClassSession objects to ICS format.

    Args:
        classes: List of class sessions
        teacher_name: Display name of the teacher for the calendar
        teacher_code: Code used in Excel (e.g. RS7); included in calendar if provided
        method: METHOD:PUBLISH (importable file) or METHOD:REQUEST (meeting request).
        invite_emails: If True, add ORGANIZER and ATTENDEE so one file can be shared with all.
        organizer_email: Email of the meeting organizer (teacher); used when invite_emails=True.

    Returns:
        ICS file content as string
    """
    calendar_title = f"{teacher_name} Schedule"
    if teacher_code and teacher_code != teacher_name:
        calendar_title = f"{teacher_name} ({teacher_code}) Schedule"

    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Building School//Timeplan//EN",
        f"X-WR-CALNAME:{_escape_ics_text(calendar_title)}",
        "CALSCALE:GREGORIAN",
        method,
    ]

    for i, session in enumerate(classes):
        # Parse start and end times
        start_hour, start_min = map(int, session.start_time.split(":"))
        end_hour, end_min = map(int, session.end_time.split(":"))

        start_dt = session.date.replace(hour=start_hour, minute=start_min, second=0)
        end_dt = session.date.replace(hour=end_hour, minute=end_min, second=0)

        # Format for ICS (local time)
        start_str = start_dt.strftime("%Y%m%dT%H%M%S")
        end_str = end_dt.strftime("%Y%m%dT%H%M%S")

        # Create unique ID (Outlook is picky: use safe chars only)
        uid = _safe_uid(f"{start_str}-{i}-{teacher_name}@buildingschool.no")

        # DTSTAMP is required by RFC 5545; Outlook requires it (UTC with Z)
        dtstamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")

        # Build summary and description from each session's own teacher_name and teacher_code
        summary = f" {session.teacher_name}"
        if session.teacher_code and session.teacher_code != session.teacher_name:
            summary += f" ({session.teacher_code})"
        summary += f" -  {session.subject} - {session.student_group}"

        description = f"Teacher: {session.teacher_name}"
        if session.teacher_code and session.teacher_code != session.teacher_name:
            description += f" ({session.teacher_code})"
        if session.room:
            description += "\nRoom: " + session.room

        event_lines = [
            "BEGIN:VEVENT",
            f"UID:{uid}",
            f"DTSTAMP:{dtstamp}",
            f"DTSTART:{start_str}",
            f"DTEND:{end_str}",
            f"SUMMARY:{_escape_ics_text(summary)}",
            f"DESCRIPTION:{_escape_ics_text(description)}",
        ]
        if session.room and session.room.strip():
            event_lines.append(f"LOCATION:{_escape_ics_text(session.room.strip())}")

        event_lines.append("END:VEVENT")
        lines.extend(event_lines)

    lines.append("END:VCALENDAR")
    return "\r\n".join(_fold_ics_line(line) for line in lines)


def print_teacher_classes_summary(
    classes: list[ClassSession], teacher_name: str, sample_size: int = 15
) -> None:
    """Print a summary of a teacher's classes (count + sample of first N)."""
    print(f"Found {len(classes)} unique classes for {teacher_name}")

    for session in classes[:sample_size]:
        print(
            f"  {session.date.strftime('%Y-%m-%d %a')} {session.start_time}-{session.end_time}: "
            f"{session.subject} ({session.student_group}) Room: {session.room or 'N/A'}"
        )

    if len(classes) > sample_size:
        print(f"  ... and {len(classes) - sample_size} more classes")


def export_teacher_calendar(
    excel_path: str,
    teacher_code: str,
    teacher_name: str | None,
    output_path: str | None = None,
) -> str:
    """Export a teacher's schedule to an ICS file.

    Args:
        excel_path: Path to the Excel timeplan file
        teacher_code: String to search for in the Excel (e.g. "RS7", "RS4")
        teacher_name: Name for the ICS/calendar; if None, defaults to teacher_code
        output_path: Optional output path for the ICS file

    Returns:
        Path to the created ICS file
    """
    if teacher_name is None:
        teacher_name = teacher_code

    classes = extract_classes_for_teacher(
        excel_path, teacher_code=teacher_code, teacher_name=teacher_name
    )
    print_teacher_classes_summary(classes, teacher_name)

    if output_path is None:
        output_path = str(Path(excel_path).parent / f"{teacher_code.lower()}_calendar.ics")

    # Main file: always PUBLISH so it can be imported
    ics_content = classes_to_ics(
        classes, teacher_name=teacher_name, teacher_code=teacher_code, method="METHOD:PUBLISH"
    )
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(ics_content)
    print(f"Calendar exported to: {output_path}")

    return output_path


def export_all_classes_of_teacher(
    excel_path: str,
    teacher_name: str,
    teacher_codes: list[str],
    output_path: str | None = None,
    sheet_name: str = "06 Timeplan",
    invite_emails: bool = False,
) -> str:
    """Export a unified ICS file for one teacher with multiple teacher codes.

    Extracts classes for each code in teacher_codes, merges and deduplicates them,
    then writes a single ICS file with teacher_name and all codes shown in the calendar.

    Args:
        excel_path: Path to the Excel timeplan file
        teacher_name: Display name for the ICS/calendar
        teacher_codes: List of strings to search for in the Excel (e.g. ["RS7", "RS4"])
        output_path: Optional output path for the ICS file
        sheet_name: Name of the sheet containing the timeplan
        invite_emails: If True, add [session.teacher_code] as ATTENDEE
            for each class session and use METHOD:REQUEST so clients can send invites.

    Returns:
        Path to the created ICS file
    """
    if not teacher_codes:
        raise ValueError("teacher_codes must contain at least one code")

    all_classes: list[ClassSession] = []
    seen_keys: set[tuple] = set()

    for code in teacher_codes:
        classes = extract_classes_for_teacher(
            excel_path,
            teacher_code=code,
            teacher_name=teacher_name,
            sheet_name=sheet_name,
        )
        for c in classes:
            key = (
                c.date.date(),
                c.start_time,
                c.end_time,
                c.subject,
                c.student_group or "Unknown",
            )
            if key in seen_keys:
                continue
            seen_keys.add(key)
            all_classes.append(c)

    all_classes = _merge_sequential_slots(all_classes)
    all_classes.sort(key=lambda c: (c.date, c.start_time))

    print_teacher_classes_summary(all_classes, teacher_name)

    if output_path is None:
        safe_name = (
            "".join(c if c.isalnum() else "_" for c in teacher_name.lower()).strip("_") or "teacher"
        )
        output_path = str(Path(excel_path).parent / f"{safe_name}_calendar.ics")

    # Use METHOD:PUBLISH so Outlook can import the file; REQUEST often fails on file import
    method = "METHOD:PUBLISH"
    organizer_email: str | None = None
    if invite_emails and teacher_codes:
        first_code = teacher_codes[0]
        first_emails = _emails_for_teacher(first_code)
        organizer_email = first_emails[0] if first_emails else None
    ics_content = classes_to_ics(
        all_classes,
        teacher_name=teacher_name,
        method=method,
        invite_emails=invite_emails,
        organizer_email=organizer_email,
    )
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(ics_content)
    print(f"Unified calendar exported to: {output_path}")

    return output_path


if __name__ == "__main__":
    """Main function to extract calendars for RS7 and RS4."""
    # excel_file = Path(__file__).parent / "01 Timeplan BYGG 2025-2026.xlsx"
    excel_file = Path("/Users/george/Downloads/1.01 Timeplan BYGG 2025-2026 (4).xlsx")

    if not excel_file.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_file}")

    print(f"Processing: {excel_file}")
    print()

    # Export GK91 calendar
    print("=" * 60)
    print("Extracting GK91 calendar")
    print("=" * 60)
    # export_teacher_calendar(str(excel_file), "RS7", teacher_name="Rand")
    export_all_classes_of_teacher(str(excel_file), "Rand", ["RS4", "RS7"], invite_emails=False)
    print()

    """# Export RS4 calendar
    print("=" * 60)
    print("Extracting RS4 calendar")
    print("=" * 60)
    export_teacher_calendar(str(excel_file), "RS4")"""
